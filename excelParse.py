import openpyxl
import datetime
import time
import re
import ipaddress
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

fileName = 'Cisco_Inventory_WorkingFile.xlsx'

# Preset for the pretty display
font = Font(bold=True)
alignment = Alignment(horizontal='center')
bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

def createWorksheet():
    
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.load_workbook(fileName)

    ws = wb['Parsed_inventory']

    ws['A2'] = nowDate
    
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'Name'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws['D4'] = 'Description'
    ws['D4'].alignment = alignment
    ws['D4'].font = font  
    ws['D4'].fill = bgColor
    ws['D4'].border = border

    ws['E4'] = 'PID'
    ws['E4'].alignment = alignment
    ws['E4'].font = font  
    ws['E4'].fill = bgColor
    ws['E4'].border = border

    ws['F4'] = 'Serial Number'
    ws['F4'].alignment = alignment
    ws['F4'].font = font  
    ws['F4'].fill = bgColor
    ws['F4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40

    wb.save(fileName)
    wb.close()

def loadExcel():

    wb = openpyxl.load_workbook(fileName)
    ws = wb['Inventory']

    excelData = list()

    for i in range(5, ws.max_row + 1):
        device = list()
        for j in range(1, ws.max_column + 1):
            cellValue = ws.cell(row = i, column = j).value
            
            if cellValue == None:
                for k in range(3, 7):
                    cellValue = ws.cell(row = i, column = k).value
                    device.append(cellValue)
                break
            else:
                device.append(cellValue)

        if len(device) != 0:
            excelData.append(device)
    
    print(excelData)
    return excelData

def saveExcelFile(deviceList):

    wb = openpyxl.load_workbook(fileName)
    ws = wb['Parsed_inventory']
    cellNumber = 5

    for elem in deviceList:
        
        # The elem contains; hostname, IP address
        if len(elem) == 2:
            ws['A' + str(cellNumber)] = elem[0]
            ws['A' + str(cellNumber)].alignment = alignment
            ws['A' + str(cellNumber)].border = border

            ws['B' + str(cellNumber)] = elem[1]
            ws['B' + str(cellNumber)].alignment = alignment
            ws['B' + str(cellNumber)].border = border

            pass

        # The elem contains; hostname, IP address, and details
        else:
            if len(elem) == 6:
                ws['A' + str(cellNumber)] = elem[0]
                ws['A' + str(cellNumber)].alignment = alignment
                ws['A' + str(cellNumber)].border = border

                ws['B' + str(cellNumber)] = elem[1]
                ws['B' + str(cellNumber)].alignment = alignment
                ws['B' + str(cellNumber)].border = border

                # The elem contains; hostname, IP address
                ws['C' + str(cellNumber)] = elem[2]
                ws['C' + str(cellNumber)].alignment = alignment
                ws['C' + str(cellNumber)].border = border

                ws['D' + str(cellNumber)] = elem[3]
                ws['D' + str(cellNumber)].alignment = alignment
                ws['D' + str(cellNumber)].border = border

                ws['E' + str(cellNumber)] = elem[4]
                ws['E' + str(cellNumber)].alignment = alignment
                ws['E' + str(cellNumber)].border = border

                ws['F' + str(cellNumber)] = elem[5]
                ws['F' + str(cellNumber)].alignment = alignment
                ws['F' + str(cellNumber)].border = border
            
            else:
                # The elem contains; details only
                ws['C' + str(cellNumber)] = elem[0]
                ws['C' + str(cellNumber)].alignment = alignment
                ws['C' + str(cellNumber)].border = border

                ws['D' + str(cellNumber)] = elem[1]
                ws['D' + str(cellNumber)].alignment = alignment
                ws['D' + str(cellNumber)].border = border

                ws['E' + str(cellNumber)] = elem[2]
                ws['E' + str(cellNumber)].alignment = alignment
                ws['E' + str(cellNumber)].border = border

                ws['F' + str(cellNumber)] = elem[3]
                ws['F' + str(cellNumber)].alignment = alignment
                ws['F' + str(cellNumber)].border = border

            cellNumber += 1

    wb.save(fileName)
    wb.close()

def filterList(string):

    if string.startswith('Te') or \
        string.startswith('Stack') or \
        string.startswith('c38xx') or \
        string.startswith('Gi') or \
        string.startswith('Chassis 1 Tran') or \
        string.startswith('Chassis 2 Tran') or \
        string.startswith('c93xx') or \
        string.startswith('Trnasceiver') or \
        string.startswith('Twen') or \
        string.startswith('Slot 1 - Tw') or \
        string.startswith('c95xx'):
        return True
    else:
        return False

def skimList(totalList):
    
    inventorySkimList = list()

    for elem in totalList:

        if len(elem) == 6:
            if filterList(elem[2]) == True:
                inventorySkimList.append(elem[0:2])
            else:
                inventorySkimList.append(elem)
        elif len(elem) == 4:
            if filterList(elem[0]) == True:
                pass
            else:
                inventorySkimList.append(elem)
    return inventorySkimList

if __name__ == '__main__':
    
    # Used already made a spreadsheet
    # Add another worksheet in existing file
    createWorksheet()
    inventoryTotalList = loadExcel()
    inventorySkimList = skimList(inventoryTotalList)
    saveExcelFile(inventorySkimList)


