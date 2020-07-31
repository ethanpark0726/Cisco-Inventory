import datetime
import time
import wexpect
import pprint
import getpass
import openpyxl
import parse
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

def createExcelFile():
    
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'
    
    # Pretty display for the File
    font = Font(bold=True)
    alignment = Alignment(horizontal='center')
    bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ws['A2'] = nowDate
    
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'Name'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws['D4'] = 'Description'
    ws['D4'].alignment = alignment
    ws['D4'].font = font  
    ws['D4'].fill = bgColor
    ws['D4'].border = border

    ws['E4'] = 'PID'
    ws['E4'].alignment = alignment
    ws['E4'].font = font  
    ws['E4'].fill = bgColor
    ws['E4'].border = border

    ws['F4'] = 'Serial Number'
    ws['F4'].alignment = alignment
    ws['F4'].font = font  
    ws['F4'].fill = bgColor
    ws['F4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40

    fileName = 'Cisco_Inventory.xlsx'
    wb.save(fileName)
    wb.close()

def saveExcelFile(deviceList, inventoryList, cellNumber):

    fileName = 'Cisco_Inventory.xlsx'
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

    ws['A' + str(cellNumber)] = deviceList[0]
    ws['A' + str(cellNumber)].alignment = alignment
    ws['A' + str(cellNumber)].border = border

    ws['B' + str(cellNumber)] = deviceList[2]
    ws['B' + str(cellNumber)].alignment = alignment
    ws['B' + str(cellNumber)].border = border

    cellIndex = cellNumber

    for i in range(0, len(inventoryList), 4):
        ws['C' + str(cellIndex)] = inventoryList[i].get('NAME')
        ws['C' + str(cellIndex)].alignment = alignment
        ws['C' + str(cellIndex)].border = border

        ws['D' + str(cellIndex)] = inventoryList[i + 1].get('DESCR')
        ws['D' + str(cellIndex)].alignment = alignment
        ws['D' + str(cellIndex)].border = border

        ws['E' + str(cellIndex)] = inventoryList[i + 2].get('PID')
        ws['E' + str(cellIndex)].alignment = alignment
        ws['E' + str(cellIndex)].border = border

        ws['F' + str(cellIndex)] = inventoryList[i + 3].get('SN')
        ws['F' + str(cellIndex)].alignment = alignment
        ws['F' + str(cellIndex)].border = border
        
        cellIndex += 1

    wb.save(fileName)

    print('--- Data successfully saved')
    wb.close()

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@is6.hsnet.ufl.edu')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    if 'SSH' in switch[3]:
        ssh_newkey = 'Are you sure you want to continue'
        session.sendline('ssh ' + switch[2])

        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline('yes')
            time.sleep(.5)
            session.sendline(password)
        elif idx == 1:
            session.sendline(password)
        
    else:
        session.sendline('telnet ' + switch[2])
        
        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect(['name', wexpect.EOF])

        if idx == 0:
            session.sendline(username)
            idx = session.expect(['word', wexpect.EOF])
            session.sendline(password)

        else:
            print('Something''s wrong!')
            print('--- Terminated Program!!')
            exit()
    idx = session.expect(['>', '#', wexpect.EOF])
    print('--- Success Login to: ', switch[2])
 
    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        
    if idx == 0:
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def getDeviceList():
    deviceList = []

    file = open('0728.txt', 'r')

    for line in file:
        temp = line.split('\t')
        temp[-1] = temp[-1].replace('\n', '')
        deviceList.append(temp)
    file.close()

    return deviceList

def commandExecute(session, os):

    command = ''

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    if os == 'IOS':
        command += 'sh inventory'
    elif os == 'NXOS':
        command += 'sh inventory'
        
    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    data = session.before.splitlines()
    return data[1:len(data) - 1]

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Cisco L2 switches Inventory Gathernig tool...            |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, Aug. 2020                        |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()

    switchList = getDeviceList()
    createExcelFile()

    for elem in switchList:
        
        session = accessJumpBox(username, password)
        session = accessSwitches(session, elem, username, password)
        data = commandExecute(session, elem[1])
        
        switch = parse.Parse(data)
        finalData = switch.getInventory()
        saveExcelFile(elem, finalData, cellNumber)

        cellNumber += len(finalData) // 4
        session.close()
